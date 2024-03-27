import openpyxl
from datetime import datetime, timedelta
import requests
import os
import json


def get_index_data(keys, year):
    words = [[{"name": keys, "wordType": 1}]]
    words = str(words).replace(" ", "").replace("'", "\"")
    startDate = f"{year}-01-01"
    endDate = f"{year}-12-31"
    url = f'http://index.baidu.com/api/SearchApi/index?area=0&word={words}&startDate={startDate}&endDate={endDate}'
    # 请求头配置
    headers = {
        "Connection": "keep-alive",
        "Accept": "application/json, text/plain, */*",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36",
        "Sec-Fetch-Site": "same-origin",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Dest": "empty",
        "Cipher-Text": "1698156005330_1698238860769_ZPrC2QTaXriysBT+5sgXcnbTX3/lW65av4zgu9uR1usPy82bArEg4m9deebXm7/O5g6QWhRxEd9/r/hqHad2WnVFVVWybHPFg3YZUUCKMTIYFeSUIn23C6HdTT1SI8mxsG5mhO4X9nnD6NGI8hF8L5/G+a5cxq+b21PADOpt/XB5eu/pWxNdwfa12krVNuYI1E8uHQ7TFIYjCzLX9MoJzPU6prjkgJtbi3v0X7WGKDJw9hwnd5Op4muW0vWKMuo7pbxUNfEW8wPRmSQjIgW0z5p7GjNpsg98rc3FtHpuhG5JFU0kZ6tHgU8+j6ekZW7+JljdyHUMwEoBOh131bGl+oIHR8vw8Ijtg8UXr0xZqcZbMEagEBzWiiKkEAfibCui59hltAgW5LG8IOtBDqp8RJkbK+IL5GcFkNaXaZfNMpI=",
        "Referer": "https://index.baidu.com/v2/main/index.html",
        "Accept-Language": "zh-CN,zh;q=0.9",
        'Cookie': Cookie}
    res = requests.get(url, headers=headers)
    res_json = res.json()

    if res_json["message"] == "bad request":
        print("抓取关键词："+keys+" 失败，请检查cookie或者关键词是否存在")
    else:
        # 获取特征值
        data = res_json['data']
        # print(data)
        uniqid = data["uniqid"]
        url = f'http://index.baidu.com/Interface/ptbk?uniqid={uniqid}'
        res = requests.get(url, headers=headers)
        # 获取解码字
        ptbk = res.json()['data']

        # 创建暂存文件夹
        os.makedirs('res', exist_ok=True)
        filename = f"{keys}_{year}.json"
        file_path = os.path.join('res', filename)
        with open(file_path, 'w', encoding='utf-8') as json_file:
            json.dump(res_json, json_file, ensure_ascii=False, indent=4)
    return file_path, ptbk

# 解码函数


def reCode(file_path, ptbk):
    # 读取暂存文件
    with open(file_path, 'r', encoding='utf-8') as file:
        res = json.load(file)
    data = res['data']
    li = data['userIndexes'][0]['all']['data']
    startDate = data['userIndexes'][0]['all']['startDate']
    year_str = startDate[:4]  # 使用切片取前四个字符，即年份部分
    try:
        # 将年份字符串转换为整数
        year = int(year_str)
        # 根据年份判断是否为闰年
        if (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0):
            year = 366
        else:
            year = 365
    except:
        year = 365

    if li == '':
        result = {}
        name = data['userIndexes'][0]['word'][0]['name']
        tep_all = []
        while len(tep_all) < year:
            tep_all.insert(0, 0)
        result["name"] = name
        result["data"] = tep_all
    else:
        ptbk = ptbk
        result = {}
        for userIndexe in data['userIndexes']:
            name = userIndexe['word'][0]['name']
            index_all = userIndexe['all']['data']
            try:
                index_all_data = [int(e) for e in decrypt(ptbk, index_all).split(",")]
                tmp_all = index_all_data
            except:
                tmp_all = []
            while len(tmp_all) < year:
                tmp_all.insert(0, 0)
            result["name"] = name
            result["data"] = tmp_all
    return result


def decrypt(ptbk, index_data):
    n = len(ptbk)//2
    a = dict(zip(ptbk[:n], ptbk[n:]))
    return "".join([a[s] for s in index_data])

# 创建日期表格


def create_excel(start_year, end_year):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # 设置第一行的标题
    sheet['A1'] = '日期'

    # 计算日期范围
    start_date = datetime(start_year, 1, 1)
    end_date = datetime(end_year, 12, 31)

    # 逐行填充日期
    current_date = start_date
    row = 2  # 从第二行开始
    while current_date <= end_date:
        sheet[f'A{row}'] = current_date.strftime('%Y-%m-%d')
        current_date += timedelta(days=1)
        row += 1

    # 保存 Excel 文件
    filename = f'百度指数数据-{start_year}-{end_year}.xlsx'
    workbook.save(filename)
    return filename

# 为文件写入数据


def write_to_excel(file_name, name, data, i):
    try:
        # 打开 Excel 文件
        workbook = openpyxl.load_workbook(file_name)
        # 获取默认的工作表（第一个工作表）
        sheet = workbook.active
        # 将名称写入第一行第i列
        sheet.cell(row=1, column=i, value=name)
        # 将数据写入从第二行开始的第i列
        for index, value in enumerate(data, start=2):
            sheet.cell(row=index, column=i, value=value)
        # 保存文件
        workbook.save(file_name)
        if len(data) != 0:
            print(f"关键词-{name}-写入成功!有效数据共{len(data)-data.count(0)}个")
    except Exception as e:
        print(f"发生错误: {e}")

# 创建日期表格


def create_excel(start_year, end_year):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # 设置第一行的标题
    sheet['A1'] = '日期'

    # 计算日期范围
    start_date = datetime(start_year, 1, 1)
    end_date = datetime(end_year, 12, 31)

    # 逐行填充日期
    current_date = start_date
    row = 2  # 从第二行开始
    while current_date <= end_date:
        sheet[f'A{row}'] = current_date.strftime('%Y-%m-%d')
        current_date += timedelta(days=1)
        row += 1

    # 保存 Excel 文件
    filename = f'百度指数数据-{start_year}-{end_year}.xlsx'
    workbook.save(filename)
    return filename


def main(keys, startDate, endDate):
    filename = create_excel(startDate, endDate)
    print(filename+"创建成功！")
    data = []
    i = 2
    for key in keys:
        for year in range(startDate, endDate + 1):
            print(f"正在处理第{year}年，请耐心等待……")
            # try:
            print(111)
            file_path = get_index_data(key, year)[0]
            print(2222)
            ptbk = get_index_data(key, year)[1]
            res = reCode(file_path, ptbk)

            name = res["name"]
            temp = res["data"]
            data = data + temp
            # except:
            #     continue
        # print(data)
        write_to_excel(filename, name, data, i)
        i = i + 1
        data = []
    print("程序运行结束！")


# 要搜索的关键词，可以输入一个列表
keys = ["开盘时间"]
# 获取的时间区间，若只获取某一年份，则二者相同
# 看起来excel处理或某个地方处理有误，例如今年是24年，取24年数据会有问题，忽略之。
# 注意！年份区间下限为2011年，不建议选择太早年份
startDate = 2011
endDate = 2023

Cookie = r'CHKFORREG=0b32e968b15fb6632785bce9cc25451b; BIDUPSID=46925A225596FBE462D97398CA155FD4; PSTM=1697729653; BAIDUID=4B11E1DFD082928EFDAE55A849E5D56D:FG=1; BAIDUID_BFESS=4B11E1DFD082928EFDAE55A849E5D56D:FG=1; ZFY=gM0chgWfq:ATfTYOcMNWTZCj:AB8w9Wzdxs:At5PR698Ps:C; H_PS_PSSID=39636_39624_39670_39664_39695_39676_39712_39764_39790_39703_39794_39681_39679; BAIDU_WISE_UID=wapp_1702402299779_203; BDUSS=hJc1ZWWkFLVzFiZW1-Z0RzbGIwbHFETW1yYnRhNmRmWWNpYVJpWml0Qkh3ZnRsSVFBQUFBJCQAAAAAAAAAAAEAAACtGhY2yOu0qM6qzfUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEc01GVHNNRldU; bdindexid=67kfrmu0ehtgu8mv6lak2mamn2; Hm_lvt_d101ea4d2a5c67dab98251f0b5de24dc=1710069526,1710133428; SIGNIN_UC=70a2711cf1d3d9b1a82d2f87d633bd8a04601458766TwRxHE4o9RGQUPI9m58NPRISuAtM%2B7h%2Bk8gHdokxS9h4ee8f6xEcfo1vkqusWAlmmrSCbP9gxDiK4beb6tH6YdpfhaGUbn4N9ZrKydfUHUk7V4KBAzjsoNwxXW3vzRYPSz9ZRMYXgcmNgMCqrGjXbfq7RQC7mxFhA6dlzDI7awGH7dS7qDfZQsexlN8pCmwMjnO5fVLK7uBjkJpNYs%2BMswJ1dNDzFFK%2FOOikrqbVdBeDEDKi1Ln5HRnsY8YWt1VLGZ26%2BOP5M76QjGAcJPmEu84aEKpO2rYGX3hUU2f3JKWCRvxhmLFwR9OPY8TYNP5G15852076490938072682080750738412; __cas__rn__=460145876; __cas__st__212=491dde0bcbc6de27188cd7cb1fe08f38826d2dcb6b88de2e28c1f7135907c507f4f4c431dce00d4f24d40b08; __cas__id__212=42346376; CPTK_212=1472132786; CPID_212=42346376; Hm_lpvt_d101ea4d2a5c67dab98251f0b5de24dc=1710134421; BDUSS_BFESS=hJc1ZWWkFLVzFiZW1-Z0RzbGIwbHFETW1yYnRhNmRmWWNpYVJpWml0Qkh3ZnRsSVFBQUFBJCQAAAAAAAAAAAEAAACtGhY2yOu0qM6qzfUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEc01GVHNNRldU; ab_sr=1.0.1_MjUzZGQ1OTk2ODNhY2E5YmM1YzNhYTMxMTdiNGFjYzMwMGZhZjUxNDFhODliZDk4YTM5ZTc1ODU5Njg2ZTQ3ZGNkZGU1ZDkyY2RmOWNkMDBkY2JjMTRjOTRjZjI0OTVkZDY5ZjFmYzJiNDFmNzAyOTkzODRhYjEyNzIzY2JmMTg1NjgzN2RhODZjMWFjMjFjNGM5NzNjNzIwOTg2MGI2NGRlNGE4MDM1ZDI1MjExMDkyNWI4ZTA3ZjlkZDA4ODg0MDIzNzgxYWM5MGE2NWY5ZGU3NjNmZDE3NmRlMmFkNmY=; RT="z=1&dm=baidu.com&si=6a089fd5-fe44-4136-9f52-415f9e99811d&ss=ltmhadpx&sl=7&tt=91t&bcn=https%3A%2F%2Ffclog.baidu.com%2Flog%2Fweirwood%3Ftype%3Dperf"'
main(keys, startDate, endDate)
